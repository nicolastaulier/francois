# -*- coding: utf-8 -*-
"""
Created on Wed Apr 19 16:13:43 2023

@author: deTRAZEGNIES
"""

import openpyxl
 
# Indiquez l’emplacement du fichier
path = r"C:/Users/deTRAZEGNIES/Desktop/Projet IHC infrasson/Code pyton spyder/Classeur1.xlsx"
 
#l'objet classeur est créé
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
 
# La boucle imprimera toutes les valeurs
# de première colonne
X = []
Y= []
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
  
    print(cell_obj.value)
    
for j in range(1, m_row + 1):
    
        cell_obj = sheet_obj.cell(row = j, column = 2)
        
        print(cell_obj.value)