# -*- coding: utf-8 -*-
"""
Created on Tue Apr 18 09:36:48 2023

@author: deTRAZEGNIES
"""

from mpl_toolkits import mplot3d
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import numpy

 
# Indiquez l’emplacement du fichier
path = r"C:/Users/deTRAZEGNIES/Desktop/Projet IHC infrasson/Code pyton spyder/Python sur excel sphères/test wxlsx.xlsx"

#l'objet classeur est créé
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
 
# La boucle imprimera toutes les valeurs
# de première colonne
R = []
rho_s= []
for i in range(2, m_row + 1):
     cell_obj = sheet_obj.cell(row = i, column = 1)
  
     rho_s.append(cell_obj.value)
rho_s = numpy.array(rho_s) 
#    print(cell_obj.value) 
    
# de deuxième colonne
    
for j in range(2, m_row + 1):
    
    cell_obj = sheet_obj.cell(row = j, column = 2)
        
    R.append(cell_obj.value)
R = numpy.array(R)
print(R) 
print(rho_s)
def func(x,y):
    rho_l = 0.998
    coeff = np.array ([1, -3*x, 0, 4*x**3*(1-y/rho_l)])
    racine = np.roots(coeff)
    # boucle appelé solution
    value=0
    for solution in racine:
        #conditions h max < 2R et solutions différentes de 0
        if solution < 2*x and solution>0:
            value = solution
    return value
    
  
# Enter les coefficients de R, rho_s et rho_l
#R=np.array ([6.25, 8.75, 23.75, 17.5, 17.5, 20.75])
#rho_s =np.array ([0.67, 0.6, 0.46, 0.41, 0.46, 0.41])
rho_l = 0.998
# ax³+bx²+cx+d -> np.roots([a,b,c,d])
"""
resultats = [] 

for i in range(0,len(R)):
    coeff = np.array ([1, -3*R[i], 0, 4*R[i]**3*(1-rho_s[i]/rho_l)])
    racine =(np.roots(coeff))
    # boucle appelé solution
    for solution in racine:
        #conditions h max < 2R et solutions différentes de 0
        if solution < 2*R[i] and solution>0:
            resultats.append(solution)
"""     
#print(resultats)

#Définir les axes du graphique x y
#x = R 
#y = np.array(resultats)

#Définir l'axe    z
N = len(R)
z=np.zeros((N,N))
z_r = np.zeros((N,N))
#print(z2)
print(z)
Z=[]
i=0
"""
for x in R:
    j=0
    for y in rho_s:
"""
for j in range(N):
        #z[i][j]= func(x,y)
        #z_r[i][j]=z[i][j]-x
        Z.append(func(R[j],rho_s[j])-R[j])

print (z)
print (z_r)


# représentation dans un graphique 3D
 
fig = plt.figure()
 
# syntaxe pour le traçage 3D
ax = plt.axes(projection ='3d')

X, Y = np.meshgrid(R, rho_s)
coinsX, coinsY = np.meshgrid([R.min(),R.max()],[rho_s.min(),rho_s.max()])
plat = np.zeros((2,2))
#Tracé d'une surface
ax.plot_surface(coinsX, coinsY, plat, cmap='winter', edgecolor='blue',alpha=0.3)
ax.scatter(R, rho_s, Z, cmap ='viridis')
# syntaxe titre en gras et axes
ax.set_title('flotabilitée fontion R, rho sphère et rho liquide', fontweight='bold')
ax.set_xlabel('Rayon sphère')
ax.set_ylabel('Masse volumique sphère')
ax.set_zlabel('Hauteur calotte-Rayon')
#affichage graphique
plt.show()
